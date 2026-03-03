"""Read input Excel files and return clean lists / lookup dicts."""

from pathlib import Path
from typing import Dict, List

import openpyxl

from src.utils import get_logger

log = get_logger(__name__)


def read_hotel_ids(xlsx_path: Path) -> List[str]:
    """Return a list of HOTELID strings from *simaFullHotelList.xlsx*.

    Uses only the column whose header is **HOTELID**.
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    # Find the HOTELID column index
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=False))
    hotel_id_col = None
    for cell in header_row:
        if cell.value and str(cell.value).strip().upper() == "HOTELID":
            hotel_id_col = cell.column
            break

    if hotel_id_col is None:
        wb.close()
        raise ValueError(f"Column 'HOTELID' not found in {xlsx_path}")

    hotel_ids: List[str] = []
    for row in ws.iter_rows(min_row=2, min_col=hotel_id_col, max_col=hotel_id_col, values_only=True):
        val = row[0]
        if val is not None:
            hotel_ids.append(str(val).strip())

    wb.close()
    log.info("Loaded %d hotel IDs from %s", len(hotel_ids), xlsx_path.name)
    return hotel_ids


def read_missing_codes(xlsx_path: Path) -> Dict[str, str]:
    """Return a dict ``{MissingSabreCode: Destination_EN}`` from the missing-codes file."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=False))
    code_col = None
    dest_col = None
    for cell in header_row:
        name = str(cell.value).strip() if cell.value else ""
        if name == "MissingSabreCode":
            code_col = cell.column
        elif name == "Destination_EN":
            dest_col = cell.column

    if code_col is None or dest_col is None:
        wb.close()
        raise ValueError(
            f"Required columns 'MissingSabreCode' / 'Destination_EN' not found in {xlsx_path}"
        )

    lookup: Dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=False):
        code_cell = row[code_col - 1]
        dest_cell = row[dest_col - 1]
        if code_cell.value is not None:
            code = str(code_cell.value).strip()
            dest = str(dest_cell.value).strip() if dest_cell.value else ""
            lookup[code] = dest

    wb.close()
    log.info("Loaded %d missing-code mappings from %s", len(lookup), xlsx_path.name)
    return lookup
